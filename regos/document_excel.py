"""
Generate Excel files for REGOS documents (purchase, wholesale, and their returns).
"""
import os
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import APP_NAME
from core.partner_terminology import get_inverted_debit_credit_labels, get_partner_document_type_name

logger = logging.getLogger(APP_NAME)


def generate_document_excel(
    document: Dict[str, Any],
    operations: List[Dict[str, Any]],
    document_type: str,
    output_dir: str = "exports"
) -> str:
    """
    Generate Excel file for a document (purchase, wholesale, or their returns).
    
    Args:
        document: Document data from REGOS API
        operations: List of operations for the document
        document_type: Type of document ("purchase", "purchase-return", "wholesale", "wholesale-return")
        output_dir: Directory to save the Excel file
        
    Returns:
        str: Path to the generated Excel file
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    
    # Determine document type label (inverted for partner view)
    type_labels = {
        "purchase": "Отгрузка",  # System purchase -> Partner sees shipment
        "purchase-return": "Возврат отгрузки",  # System purchase return -> Partner sees shipment return
        "wholesale": "Закупка",  # System wholesale -> Partner sees purchase
        "wholesale-return": "Возврат закупки"  # System wholesale return -> Partner sees purchase return
    }
    doc_type_label = type_labels.get(document_type, "Документ")
    
    # Determine if we use cost or price
    use_cost = document_type in ["purchase", "purchase-return"]
    is_return = document_type in ["purchase-return", "wholesale-return"]
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Document header
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = f"🧾 {doc_type_label}"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 2
    
    # Document information
    doc_code = document.get("code", "N/A")
    doc_date = document.get("date", "")
    
    # Format date
    if isinstance(doc_date, (int, float)):
        formatted_date = datetime.fromtimestamp(doc_date).strftime("%d.%m.%Y %H:%M")
    else:
        try:
            formatted_date = datetime.fromisoformat(str(doc_date)).strftime("%d.%m.%Y %H:%M")
        except:
            formatted_date = str(doc_date)
    
    info_rows = [
        ("Номер документа:", doc_code),
        ("Дата:", formatted_date),
    ]
    
    # Add warehouse if available
    stock = document.get("stock")
    if stock:
        stock_name = stock.get("name", "") if isinstance(stock, dict) else str(stock)
        if stock_name:
            info_rows.append(("Склад:", stock_name))
    
    for label, value in info_rows:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = title_font
        ws.cell(row=row, column=2).value = value
        row += 1
    
    row += 1
    
    # Operations header
    headers = ["№", "Товар", "Код", "Штрихкод", "Количество", "Цена/Стоимость", "Сумма"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[row].height = 20
    row += 1
    
    # Add operations
    total_items = 0
    total_amount = 0.0
    
    for idx, operation in enumerate(operations, 1):
        item = operation.get("item", {})
        if isinstance(item, dict):
            item_name = item.get("name", "Неизвестный товар")
            item_code = item.get("code", "")
            item_barcode = item.get("base_barcode", "")
        else:
            item_name = str(item) if item else "Неизвестный товар"
            item_code = ""
            item_barcode = ""
        
        quantity = float(operation.get("quantity", 0))
        
        # Use cost for purchase, price for wholesale
        if use_cost:
            price = float(operation.get("cost", 0))
        else:
            price = float(operation.get("price", 0))
        
        item_total = quantity * price
        total_items += quantity
        total_amount += item_total
        
        # Add row data
        row_data = [
            idx,
            item_name,
            item_code,
            item_barcode,
            quantity,
            price,
            item_total
        ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col)
            if col == 1:
                cell.value = value  # Number
            elif col in [2, 3, 4]:
                cell.value = value  # Item name, code, barcode (text)
            else:
                cell.value = value  # Numeric values
                cell.number_format = '#,##0.00'
            cell.border = border
            cell.alignment = Alignment(
                horizontal='left' if col in [2, 3, 4] else 'right',
                vertical='center'
            )
        
        row += 1
    
    # Add totals
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "Всего товаров:"
    cell.font = bold_font
    cell.alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=5).value = total_items
    ws.cell(row=row, column=5).font = bold_font
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=5).border = border
    
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "Итого к оплате:"
    cell.font = bold_font
    cell.font = Font(bold=True, size=12, color="0088CC")
    cell.alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=6).value = total_amount
    ws.cell(row=row, column=6).font = Font(bold=True, size=12, color="0088CC")
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=6).border = border
    
    # Adjust column widths
    column_widths = [8, 40, 15, 18, 15, 18, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Generate filename - reverse the document type prefix
    # purchase -> wholesale, wholesale -> purchase
    # purchase-return -> wholesale-return, wholesale-return -> purchase-return
    filename_prefix_map = {
        "purchase": "wholesale",
        "wholesale": "purchase",
        "purchase-return": "wholesale-return",
        "wholesale-return": "purchase-return"
    }
    reversed_type = filename_prefix_map.get(document_type, document_type)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    doc_id = document.get("id", "unknown")
    filename = f"{reversed_type}_{doc_id}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Save workbook
    wb.save(filepath)
    logger.info(f"Generated Excel file: {filepath}")
    
    return filepath


def generate_partner_balance_excel(
    balance_entries: List[Dict[str, Any]],
    output_dir: str = "exports"
) -> str:
    """
    Generate Excel file for partner balance with totals.
    Groups entries by currency and firm, shows totals.
    
    Args:
        balance_entries: List of balance entries from PartnerBalance/Get
        output_dir: Directory to save the Excel file
        
    Returns:
        str: Path to the generated Excel file
    """
    if not balance_entries:
        raise ValueError("No balance data to export")
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Group entries by currency and firm
    from collections import defaultdict
    grouped_data = defaultdict(lambda: defaultdict(list))
    
    for entry in balance_entries:
        currency = entry.get("currency", {})
        firm = entry.get("firm", {})
        
        currency_name = currency.get("name", "Unknown") if isinstance(currency, dict) else "Unknown"
        firm_name = firm.get("name", "Unknown") if isinstance(firm, dict) else "Unknown"
        
        grouped_data[currency_name][firm_name].append(entry)
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    total_font = Font(bold=True, size=11, color="008000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create a sheet for each currency
    for currency_name, firms_data in grouped_data.items():
        # Create sheet for currency
        sheet_name = currency_name[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        
        row = 1
        
        # Currency header
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"💱 Валюта: {currency_name}"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 25
        row += 2
        
        # Process each firm
        currency_total_debit = 0
        currency_total_credit = 0
        currency_total_start = 0
        
        for firm_name, entries in firms_data.items():
            # Firm header
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws[f'A{row}']
            cell.value = f"🏢 Предприятие: {firm_name}"
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='left')
            row += 1
            
            # Column headers (inverted for partner view)
            partner_debit_label, partner_credit_label = get_inverted_debit_credit_labels("ru")
            headers = ["Дата", "Документ", "Тип документа", "Начальный остаток", partner_debit_label, partner_credit_label, "Остаток", "Курс"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            ws.row_dimensions[row].height = 20
            row += 1
            
            # Sort entries by date
            entries.sort(key=lambda x: x.get("date", 0))
            
            # Add entries
            firm_total_debit = 0
            firm_total_credit = 0
            # Start amount is from the first entry (oldest)
            firm_total_start = float(entries[0].get("start_amount", 0)) if entries else 0
            
            for entry in entries:
                entry_date = entry.get("date", 0)
                doc_code = entry.get("document_code", "N/A")
                doc_type = entry.get("document_type", {})
                doc_type_name_raw = doc_type.get("name", "Неизвестно") if isinstance(doc_type, dict) else "Неизвестно"
                # Convert to partner perspective
                doc_type_name = get_partner_document_type_name(doc_type_name_raw, "ru")
                start_amount = float(entry.get("start_amount", 0))
                debit = float(entry.get("debit", 0))
                credit = float(entry.get("credit", 0))
                remainder = start_amount + debit - credit
                exchange_rate = entry.get("exchange_rate", 1.0)
                
                # Format date
                if isinstance(entry_date, (int, float)):
                    formatted_date = datetime.fromtimestamp(entry_date).strftime("%d.%m.%Y %H:%M")
                else:
                    formatted_date = str(entry_date)
                
                # Swap debit/credit values for partner view (inverted terminology)
                # System debit -> Partner credit column, System credit -> Partner debit column
                row_data = [
                    formatted_date,
                    doc_code,
                    doc_type_name,
                    start_amount,
                    credit if credit != 0 else None,  # System credit -> Partner debit column
                    debit if debit != 0 else None,     # System debit -> Partner credit column
                    remainder,
                    exchange_rate if exchange_rate != 1.0 else None
                ]
                
                for col, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col)
                    if value is None:
                        cell.value = "—"
                    elif isinstance(value, (int, float)):
                        cell.value = value
                        cell.number_format = '#,##0.00'
                    else:
                        cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left' if col in [1, 2, 3] else 'right')
                
                firm_total_debit += debit
                firm_total_credit += credit
                row += 1
            
            # Firm totals row
            # Calculate final remainder: initial start + all debits - all credits
            # The last entry's remainder is the final balance
            last_entry = entries[-1] if entries else None
            if last_entry:
                firm_remainder = float(last_entry.get("start_amount", 0)) + float(last_entry.get("debit", 0)) - float(last_entry.get("credit", 0))
            else:
                firm_remainder = firm_total_start + firm_total_debit - firm_total_credit
            row += 1
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = f"Итого ({firm_name}):"
            cell.font = bold_font
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            
            ws.cell(row=row, column=4).value = firm_total_start
            ws.cell(row=row, column=4).font = bold_font
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=4).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            
            # Swap debit/credit totals for partner view (inverted terminology)
            ws.cell(row=row, column=5).value = firm_total_credit if firm_total_credit != 0 else None  # System credit -> Partner debit column
            ws.cell(row=row, column=5).font = bold_font
            if firm_total_credit != 0:
                ws.cell(row=row, column=5).number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=5).value = "—"
            ws.cell(row=row, column=5).border = border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            
            ws.cell(row=row, column=6).value = firm_total_debit if firm_total_debit != 0 else None  # System debit -> Partner credit column
            ws.cell(row=row, column=6).font = bold_font
            if firm_total_debit != 0:
                ws.cell(row=row, column=6).number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=6).value = "—"
            ws.cell(row=row, column=6).border = border
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            
            ws.cell(row=row, column=7).value = firm_remainder
            ws.cell(row=row, column=7).font = total_font
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=7).border = border
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=7).fill = PatternFill(start_color="E7F4E4", end_color="E7F4E4", fill_type="solid")
            
            row += 2
            
            # For currency totals, track the initial start amount from the first firm's first entry
            if currency_total_start == 0 and entries:
                currency_total_start = float(entries[0].get("start_amount", 0))
            currency_total_debit += firm_total_debit
            currency_total_credit += firm_total_credit
        
        # Currency totals row
        # For currency totals, calculate based on all entries across all firms
        # Get the last entry across all firms for this currency to get final remainder
        all_currency_entries = []
        for firm_entries in firms_data.values():
            all_currency_entries.extend(firm_entries)
        all_currency_entries.sort(key=lambda x: x.get("date", 0))
        
        if all_currency_entries:
            last_currency_entry = all_currency_entries[-1]
            currency_remainder = float(last_currency_entry.get("start_amount", 0)) + float(last_currency_entry.get("debit", 0)) - float(last_currency_entry.get("credit", 0))
        else:
            currency_remainder = currency_total_start + currency_total_debit - currency_total_credit
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = f"ВСЕГО ({currency_name}):"
        cell.font = Font(bold=True, size=13, color="0000FF")
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        cell.fill = PatternFill(start_color="D0E8F2", end_color="D0E8F2", fill_type="solid")
        
        ws.cell(row=row, column=4).value = currency_total_start
        ws.cell(row=row, column=4).font = Font(bold=True, size=13, color="0000FF")
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=4).fill = PatternFill(start_color="D0E8F2", end_color="D0E8F2", fill_type="solid")
        
        # Swap debit/credit totals for partner view (inverted terminology)
        ws.cell(row=row, column=5).value = currency_total_credit if currency_total_credit != 0 else None  # System credit -> Partner debit column
        ws.cell(row=row, column=5).font = Font(bold=True, size=13, color="0000FF")
        if currency_total_credit != 0:
            ws.cell(row=row, column=5).number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=5).value = "—"
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5).fill = PatternFill(start_color="D0E8F2", end_color="D0E8F2", fill_type="solid")
        
        ws.cell(row=row, column=6).value = currency_total_debit if currency_total_debit != 0 else None  # System debit -> Partner credit column
        ws.cell(row=row, column=6).font = Font(bold=True, size=13, color="0000FF")
        if currency_total_debit != 0:
            ws.cell(row=row, column=6).number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=6).value = "—"
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6).fill = PatternFill(start_color="D0E8F2", end_color="D0E8F2", fill_type="solid")
        
        ws.cell(row=row, column=7).value = currency_remainder
        ws.cell(row=row, column=7).font = Font(bold=True, size=13, color="0000FF")
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        ws.cell(row=row, column=7).border = border
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=7).fill = PatternFill(start_color="D0E8F2", end_color="D0E8F2", fill_type="solid")
        
        # Adjust column widths
        column_widths = [18, 15, 30, 18, 15, 15, 18, 12]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"partner_balance_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Save workbook
    wb.save(filepath)
    logger.info(f"Generated partner balance Excel file: {filepath}")
    
    return filepath
